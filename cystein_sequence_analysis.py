import re

from pathlib import Path
from Bio import AlignIO, SeqIO
from collections import Counter
import numpy as np
import pandas as pd # (+ install tabulate)
import matplotlib.pyplot as plt
from matplotlib.patches import Patch
import logomaker
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from sklearn.cluster import AgglomerativeClustering
from sklearn.metrics import silhouette_score, adjusted_rand_score
from scipy.cluster.hierarchy import linkage, dendrogram, fcluster, cophenet
from scipy.spatial.distance import pdist
from scipy.spatial.distance import squareform
from typing import Dict, Optional, Tuple



def process_alignment_DSATYY_WGXG(aln_path: Path, output_fasta: Path, motif: str = "DSATYY"
) -> None:
    """
    1) Load a precomputed Clustal alignment (no re-alignment performed)
    2) Identify the motif start position in each aligned sequence
    3) print distribution of positions
    4) determine most frequent position
    5) For each sequence: advance from the consensus start position while
       consuming exactly the number of amino acids corresponding to the motif
       length, explicitly skipping alignment gaps ('-') to ensure complete
       motif consumption
    6) Truncate each sequence immediately after the fully consumed DSATYY motif
    7) remove gaps ('-')
    8) cut last 4 residues (WGXG)
    9) write resulting sequences to FASTA
    """

    # --- load alignment ---
    alignment = AlignIO.read(aln_path, "clustal")

    motif_positions = {}
    position_counter = Counter()
    n_with_motif = 0

    # --- find motif position per sequence ---
    for record in alignment:
        seq_aln = str(record.seq)
        pos = seq_aln.find(motif)

        if pos != -1:
            position_counter[pos] += 1
            n_with_motif += 1

    # --- print position distribution ---
    if not position_counter:
        raise ValueError(f"{motif} motif not found in any sequence.")
    print(
        f"{motif} positions (pos:count): "
        + ", ".join(f"{pos}:{count}" for pos, count in sorted(position_counter.items()))
        + f" | coverage: {n_with_motif}/{len(alignment)} sequences"
    )

    # --- process sequences ---
    processed_records = []
    consensus_pos = position_counter.most_common(1)[0][0]

    for record in alignment:
        seq_aln = str(record.seq)

        if consensus_pos >= len(seq_aln):
            continue

        # --- CUT LOGIC: consume len(motif) amino acids, skipping gaps ---
        cut_idx = consensus_pos
        aa_consumed = 0

        while cut_idx < len(seq_aln) and aa_consumed < len(motif):
            if seq_aln[cut_idx] != "-":
                aa_consumed += 1
            cut_idx += 1

        if aa_consumed < len(motif):
            continue  # motif incomplete, skip sequence

        # cut sequence AFTER motif (biologically correct)
        seq_cut = seq_aln[cut_idx:]

        # remove gaps
        seq_cut = seq_cut.replace("-", "")

        # remove last 4 residues (WGXG tail)
        if len(seq_cut) > 4:
            seq_cut = seq_cut[:-4]
            processed_records.append((record.id, seq_cut))

    # --- write FASTA ---
    with open(output_fasta, "w") as f:
        for i, (rid, seq) in enumerate(processed_records, start=1):
            f.write(f">seq_{i}\n")
            f.write(f"{seq}\n")

    print(f"Cut {motif} at most frequent position ({consensus_pos}) and last four aa and written {len(processed_records)} sequences to: {output_fasta}")


def export_gapfree_sequences_from_alignment(aln_path: Path, output_fasta: Path
) -> None:
    """
    Load an existing Clustal alignment, remove all gap characters ('-'),
    and write the resulting gap-free sequences to a FASTA file.
    """

    # load alignment
    alignment = AlignIO.read(aln_path, "clustal")

    processed = []

    for record in alignment:
        seq = str(record.seq).replace("-", "")
        if seq:  # skip empty sequences
            processed.append(seq)

    if not processed:
        raise ValueError("No gap-free sequences produced from alignment.")

    # write FASTA
    with open(output_fasta, "w") as f:
        for i, seq in enumerate(processed, start=1):
            f.write(f">knob_{i}\n")
            f.write(f"{seq}\n")

    print(f"Written {len(processed)} gap-free {aln_path.stem} sequences to: {output_fasta}")


def process_alignment_by_conserved_c_positions(aln_path: Path, output_dir: Path, output_suffix: str, n_ignore: int = 32, c_threshold: float = 0.5
) -> Path:
    """
    Process an alignment using conserved cysteine positions
    (ignoring the first n_ignore sequences).

     Returns:
        Path to gap-free FASTA file with processed sequences

    1. Loads an existing Clustal alignment
    2. Completely ignores the first `n_ignore` sequences (default: 32)
    3. Works only on the remaining sequences
    4. Finds the most frequent common C positions
    5. Prints their position distribution
    6. Selects:
        • the most frequent C position
        • the second most frequent C position
        • (optionally controllable via threshold)
    7. Determines the difference between these two positions
    8. Counts individually per sequence the number of non-gap AS between these two positions
    9. Trims:
        • left: directly before the second-most-frequent C position
        • left part is discarded
        • right: additionally exactly as many AS as were previously counted between the two C positions
    10. Removes all gaps
    11. Writes a new FASTA
    """

    alignment = AlignIO.read(aln_path, "clustal")

    if len(alignment) <= n_ignore:
        raise ValueError("Alignment contains fewer sequences than n_ignore.")

    # ---- sequences of interest (ignore first n_ignore) ----
    target_seqs = alignment[n_ignore:]
    n_target = len(target_seqs)
    aln_len = alignment.get_alignment_length()

    # ---- count cysteine positions ----
    c_position_counter = Counter()

    for pos in range(aln_len): # is sequence length, not sequence amount
        column = target_seqs[:, pos]
        c_count = column.count("C")

        if (c_count / n_target) >= c_threshold:
            c_position_counter[pos] = c_count

    # ---- print distribution ----
    print(
        "Conserved C positions (pos:count): "
        + ", ".join(f"{p}:{c}" for p, c in c_position_counter.items()))

    # ---- select two most frequent C positions ----
    if len(c_position_counter) < 2:
        raise ValueError("Fewer than two conserved C positions found.")
    sorted_c_positions = sorted(c_position_counter.keys()) # sorted via keys (position)
    c_left = sorted_c_positions[0]
    c_right = sorted_c_positions[1]

    delta = c_right - c_left
    print(f"Selected leftmost C positions: {c_left}, {c_right} (diff: {delta})")

    # ---- process sequences ----
    processed = []

    for record in target_seqs:
        seq_aln = str(record.seq)

        # count AA between the two C positions (exclude gaps)
        aa_between = sum(
            1 for aa in seq_aln[c_left + 1: c_right] if aa != "-"
        )

        # cut RIGHT part: everything AFTER second C
        right_part = seq_aln[c_right:]

        # remove gaps
        ungapped = right_part.replace("-", "")

        # subtract individual aa count from the END
        if aa_between >= len(ungapped):
            continue

        if aa_between == 0:
            final_seq = ungapped
        else:
            final_seq = ungapped[:-aa_between]
        processed.append(final_seq)

    # ---- write FASTA ----
    output_dir.mkdir(exist_ok=True)
    fasta_out = output_dir / f"cut_knobs_{output_suffix}.fasta"

    with open(fasta_out, "w") as f:
        for i, seq in enumerate(processed, start=1):
            f.write(f">seq_{i}\n{seq}\n")

    print(f"Written {len(processed)} processed {aln_path.stem} sequences to: {fasta_out}")

    return fasta_out


def levenshtein_distance(seq1: str, seq2: str) -> int:
    """
    Compute edit distance between two sequences:
    minimal number of operations (insertions, deletions, substitutions) to transform one into the other.
    """
    n, m = len(seq1), len(seq2)

    dp = [[0] * (m + 1) for _ in range(n + 1)] # creates matrix wth n+1 rows and m+1 columns

    for i in range(n + 1): # writes first column (if only deletions)
        dp[i][0] = i
    for j in range(m + 1): # writes first row (if only deletions)
        dp[0][j] = j

    for i in range(1, n + 1): # iterates over remaining rows
        for j in range(1, m + 1): # iterates over remaining columns
            cost = 0 if seq1[i - 1] == seq2[j - 1] else 1 # 0 if equal letter, 1 if different
            dp[i][j] = min(
                dp[i - 1][j] + 1,      # deletion (from above: i-1, j)
                dp[i][j - 1] + 1,      # insertion (from left: i, j-1)
                dp[i - 1][j - 1] + cost  # substitution (from diagonal: i-1, j-1)
            )

    return dp[n][m] # gives nack last entry of last row (minimal distance)


def deduplicate_and_filter_fasta(input_fasta: Path, output_fasta: Path, max_aa_difference: int = 0, write_output: bool = True,
                                 negative_binder_ids: Optional[set[str]] = None, debug: bool = False,
) -> dict:
    """
    Remove exact and near-duplicate sequences (unaligned FASTA).
    Uses Levenshtein distance for robustness.
    """
    seen_sequences = {}
    removed_records = []

    records = list(SeqIO.parse(input_fasta, "fasta"))

    if debug:
        print(f"\n=== Deduplication START ===")
        print(f"Input FASTA: {input_fasta}")
        print(f"Total sequences: {len(records)}")
        print(f"Max AA difference: {max_aa_difference}")

    # STEP 1: EXACT DEDUP
    temp_records = []
    for record in records:
        seq_str = str(record.seq)

        if seq_str not in seen_sequences or (negative_binder_ids and record.id in negative_binder_ids):
            seen_sequences[seq_str] = record.id
            temp_records.append(record)
        else:
            removed_records.append({
                "removed_id": record.id,
                "kept_id": seen_sequences[seq_str],
                "sequence": seq_str,
                "reason": "exact_duplicate",})

            if debug:
                print(f"[EXACT DUP] removed {record.id} "
                      f"(duplicate of {seen_sequences[seq_str]})")

    if debug:
        print(f"\nAfter exact dedup: {len(temp_records)} sequences")

    # STEP 2: NEAR-DUPLICATES
    final_records = []
    kept_sequences = []

    for record in temp_records:
        seq_str = str(record.seq)

        is_duplicate = False

        # negative binders are kept
        if negative_binder_ids and record.id in negative_binder_ids:
            final_records.append(record)
            kept_sequences.append((seq_str, record.id))
            if debug:
                print(f"[KEEP NEG] {record.id}")
            continue

        for kept_seq, kept_id in kept_sequences:
            dist = levenshtein_distance(seq_str, kept_seq)

            if debug:
                print(
                    f"[COMPARE] {record.id} vs {kept_id} → dist={dist}")

            if dist <= max_aa_difference:
                removed_records.append({
                    "removed_id": record.id,
                    "kept_id": kept_id,
                    "sequence": seq_str,
                    "reason": f"edit_distance <= {max_aa_difference}"})

                if debug:
                    print(
                        f"[NEAR DUP] removed {record.id} "
                        f"(similar to {kept_id}, dist={dist})")

                is_duplicate = True
                break

        if not is_duplicate:
            final_records.append(record)
            kept_sequences.append((seq_str, record.id))

            if debug:
                print(f"[KEEP] {record.id}")

    # WRITE OUTPUT
    if write_output:
        with open(output_fasta, "w") as handle:
            SeqIO.write(final_records, handle, "fasta")

        print(f"Input FASTA          : {input_fasta}")
        print(f"Output FASTA         : {output_fasta}")
        print(f"Sequences kept       : {len(final_records)}")
        print(f"Total removed        : {len(removed_records)}")

        n_exact = sum(r["reason"] == "exact_duplicate" for r in removed_records)
        print(f"Exact duplicates     : {n_exact}")
        print(f"Near duplicates      : {len(removed_records) - n_exact}")

    # optional full debug dump
    if debug and removed_records:
        print("\n--- Removed sequences (summary) ---")
        for r in removed_records:
            print(
                f"  removed {r['removed_id']} "
                f"(kept {r['kept_id']}, reason: {r['reason']})")
    if debug:
        print("\n=== Deduplication END ===")

    return {
        "kept_ids": [rec.id for rec in final_records],
        "removed": removed_records,}


def evaluate_dedup_thresholds(input_fasta: Path, diff_range: range):
    """
    Evaluate different max_aa_difference thresholds and print statistics.
    """
    results = []  # store results instead of printing immediately
    for diff in diff_range:
        result = deduplicate_and_filter_fasta(
            input_fasta=input_fasta,
            output_fasta=Path("dummy.fasta"),
            max_aa_difference=diff,
            write_output=False)

        removed = result["removed"]

        exact = sum(r["reason"] == "exact_duplicate" for r in removed)
        near = len(removed) - exact
        kept = len(result["kept_ids"])

        results.append((diff, exact, near, kept))

    # --- print once ---
    print("\n=== DEDUPLICATION THRESHOLD EVALUATION ===")
    print(f"Input FASTA: {input_fasta}")
    print(f"{'max_diff':>8} | {'exact':>6} | {'near':>6} | {'kept':>6}")
    print("-" * 40)

    for diff, exact, near, kept in results:
        print(f"{diff:>8} | {exact:>6} | {near:>6} | {kept:>6}")
    print("Recommendation: choose based on diversity vs redundancy trade-off:")


def cysteine_clustering(aln_path: Path, output_dir: Path, output_prefix: str, n_ignore: int = 0,
                        cluster_range: range = range(2, 10), highlight_mode: str = "none", highlight_n: int = 0,
                        negative_binder_ids: Optional[Dict[str, Dict]] = None, near_knob_similarity_factor: float = 0.4,
                        debug: bool = False
                        ) -> tuple[dict[int, dict], np.ndarray, list[str], dict[str, list[str]], set[str]]:
    """
    Perform hierarchical clustering of sequences based on cysteine (C) patterns.
    Method:
    1. Load Clustal alignment
    2. Ignore the first n_ignore sequences (reference / literature)
    3. Encode sequences as binary vectors (1 = C, 0 = non-C)
    4. Perform hierarchical clustering (Manhattan distance, average linkage)
    5. Select optimal cluster number using silhouette score
    6. Visualize dendrogram of final clustering: Sequence names and Background in cluster color,
                                                 Mark all structure knobs and all near ones

    Returns
    -------
    Path
       Path to the dendrogram figure, Cluster data structure, Clustering labels, and Sequence IDs.
    """

    # --- load alignment & select sequences ---
    alignment = AlignIO.read(aln_path, "clustal")
    if len(alignment) <= n_ignore:
        raise ValueError("Alignment contains fewer sequences than n_ignore.")

    target_records = alignment[n_ignore:]
    ids = [rec.id for rec in target_records]

    if debug:
        print("\n=== Cysteine clustering START ===")
        print(f"Alignment: {aln_path}")
        print(f"Total sequences: {len(alignment)}")
        print(f"Ignoring first n={n_ignore}")
        print(f"Target sequences used: {len(target_records)}")
        print(f"Example IDs: {ids[:5]}")

    # --- define knob sequences ---
    if highlight_mode == "first_n" and highlight_n > 0:
        highlight_ids = {rec.id for rec in alignment[:highlight_n]}
    else:
        highlight_ids = set()
    knob_ids = set(highlight_ids)

    # --- prepare negative binder IDs ---
    if negative_binder_ids is not None:
        negative_binder_id_set = set(negative_binder_ids.keys())
        missing_negative_ids = negative_binder_id_set - set(ids)
        if missing_negative_ids:
            print(f"Negative binder IDs not found in dendrogram: {', '.join(sorted(missing_negative_ids))}")
        if debug:
            print(f"Negative binders in data: {len(negative_binder_id_set)}")
    else:
        negative_binder_id_set = set()

    # --- encode cysteine pattern (binary matrix) ---
    patterns = np.array([
        [1 if aa == "C" else 0 for aa in str(rec.seq)]
        for rec in target_records])
    if debug:
        print("\nCysteine pattern matrix:")
        print(f"Shape: {patterns.shape}")
    if debug and len(patterns) > 0:
        print(f"Example sequence: {target_records[0].id}")
        print(f"Pattern: {patterns[0]}")

    # --- find optimal number of clusters ---
    best_score = -1.0
    best_labels = None
    best_k = None

    for k in cluster_range:
        model = AgglomerativeClustering(
            n_clusters=k,
            linkage="average",
            metric="manhattan",
        )
        labels = model.fit_predict(patterns)
        score = silhouette_score(patterns, labels, metric="manhattan")

        if debug:
            print(f"[CLUSTER TEST] k={k}, silhouette={score:.3f}")

        if score > best_score:
            best_score = score
            best_k = k
            best_labels = labels

    # --- build cluster data structure ---
    clusters = {}
    for idx, (seq_id, label, record) in enumerate(
        zip(ids, best_labels, target_records)):
        clusters.setdefault(label, {
            "indices": [],
            "ids": [],
            "sequences": [],})
        clusters[label]["indices"].append(idx)
        clusters[label]["ids"].append(seq_id)
        clusters[label]["sequences"].append(str(record.seq))

    # --- assign global display ids ---
    global_display_counter = 1

    for cluster_id in sorted(clusters):

        seq_ids = [str(s) for s in clusters[cluster_id]["ids"]]

        def extract_number(s):
            match = re.search(r"(\d+)", s)
            return int(match.group(1)) if match else float("inf")

        seq_ids = sorted(seq_ids, key=extract_number)

        display_ids = {}
        for seq_id in seq_ids:
            display_ids[seq_id] = global_display_counter
            global_display_counter += 1
        clusters[cluster_id]["display_ids"] = display_ids

    if debug:
        print("\nCluster summary:")
        for label, data in clusters.items():
            print(
                f"Cluster {label}: "
                f"{len(data['ids'])} sequences")

    # --- hierarchical linkage (SciPy) ---
    Z = linkage(patterns, method="average", metric="cityblock")

    labels_scipy = fcluster(Z, t=best_k, criterion="maxclust")
    ari = adjusted_rand_score(labels_scipy, best_labels)
    print(
        f"Best cluster number: {best_k} "
        f"(silhouette score: {best_score:.3f}), "
        f"Similarity sklearn vs. SciPy clustering (ARI): {ari:.3f}")

    # --- dendrogram plot ---
    plt.figure(figsize=(20, 10))
    plt.subplots_adjust(bottom=0.1)

    id_to_cluster = dict(zip(ids, best_labels))

    # --- cluster colors ---
    cluster_colors = [
        "tab:blue", "tab:orange", "tab:green", "tab:red",
        "tab:purple", "tab:brown", "tab:pink", "tab:gray", "tab:olive",

        # additional distinct colors
        "cyan", "magenta", "gold", "lime", "teal", "navy",
        "maroon", "darkorange", "darkgreen", "indigo", "crimson",]

    unique_clusters = sorted(set(best_labels))
    cluster_to_color = {
        cluster: cluster_colors[i]
        for i, cluster in enumerate(unique_clusters)}

    distance_threshold = Z[-best_k + 1, 2]
    ddata = dendrogram(
        Z,
        labels=ids,
        leaf_rotation=90,
        leaf_font_size=4,
        color_threshold=0,
        above_threshold_color="black")
    ax = plt.gca()

    # --- BACKGROUND SHADING FOR CLUSTERS ---
    leaf_order = ddata["leaves"]
    # get cluster per leaf in plot order
    ordered_clusters = [best_labels[i] for i in leaf_order]
    # find continuous cluster blocks
    blocks = []
    start = 0
    for i in range(1, len(ordered_clusters)):
        if ordered_clusters[i] != ordered_clusters[i - 1]:
            blocks.append((start, i - 1, ordered_clusters[i - 1]))
            start = i
    blocks.append((start, len(ordered_clusters) - 1, ordered_clusters[-1]))

    # plot shaded regions
    for start, end, cluster in blocks:
        x_start = start * 10
        x_end = (end + 1) * 10

        ax.axvspan(
            x_start,
            x_end,
            color=cluster_to_color[cluster],
            alpha=0.08,  # transparency
            zorder=0)  # keep behind dendrogram

    plt.title(
        "Hierarchical clustering based on cysteine patterns",
        fontsize=22,
        fontweight="bold")
    plt.xlabel("Sequences", fontsize=16)
    plt.ylabel("Manhattan distance", fontsize=16)

    # --- compute neighbors of knobs ---
    id_to_index = {seq_id: i for i, seq_id in enumerate(ids)}
    knob_indices = [
        id_to_index[sid] for sid in knob_ids if sid in id_to_index]

    _, coph_condensed = cophenet(Z, pdist(patterns))
    coph_dists = squareform(coph_condensed)
    neighbor_distance_threshold = distance_threshold * near_knob_similarity_factor

    if debug:
        print("\nDistance threshold:")
        print(f"Raw threshold: {distance_threshold:.3f}")
        print(f"Adjusted (factor={near_knob_similarity_factor}): {neighbor_distance_threshold:.3f}")

    near_knob_ids = set()
    for i in knob_indices:
        for j, seq_id in enumerate(ids):
            if i != j and coph_dists[i, j] <= neighbor_distance_threshold:
                near_knob_ids.add(seq_id)
    near_knob_ids -= knob_ids

    if debug:
        print(f"\nNear-knobs identified: {len(near_knob_ids)}")
        print(f"Example: {list(near_knob_ids)[:5]}")

    near_knob_to_knobs = {}
    for i in knob_indices:
        knob_id = ids[i]

        for j, seq_id in enumerate(ids):
            if i != j and coph_dists[i, j] <= neighbor_distance_threshold:

                if seq_id in knob_ids:
                    continue # no knob to knob mapping
                near_knob_to_knobs.setdefault(seq_id, []).append(knob_id)
    near_knob_to_knobs = {
        k: v for k, v in near_knob_to_knobs.items()
        if k not in knob_ids}

    if debug:
        print("\nNear-knob relationships:")
        for seq, knobs in near_knob_to_knobs.items():
            print(f"{seq} -> {', '.join(knobs)}")

    # --- color dendrogram ---
    for label, leaf_idx in zip(ax.get_xmajorticklabels(), ddata["leaves"]):
        seq_id = ids[leaf_idx]

        # knobs
        if seq_id in knob_ids:
            label.set_fontweight("bold")
            label.set_bbox(
                dict(facecolor="lightgray", edgecolor="none", boxstyle="round,pad=0.15"))
        # negative
        elif seq_id in negative_binder_id_set:
            label.set_bbox(
                dict(facecolor="lightcoral", edgecolor="none", boxstyle="round,pad=0.15"))
        elif seq_id in near_knob_ids:
            label.set_bbox(
                dict(facecolor="orange", edgecolor="none", boxstyle="round,pad=0.15"))

    ax.set_ylim(0, Z[:, 2].max() * 1.05)

    # --- legend ---
    legend_elements = [
        Patch(
            facecolor=cluster_to_color[c],
            label=f"Cluster {c + 1} (n={len(clusters[c]['sequences'])})"
        )
        for c in sorted(unique_clusters)]

    highlight_legend = [
        Patch(facecolor="lightgray", edgecolor="black", label="Knobs"),
        Patch(facecolor="orange", edgecolor="black", label="Near knobs")]
    if negative_binder_id_set:
        highlight_legend = [
            Patch(facecolor="lightgray", edgecolor="black", label="Knobs"),
            Patch(facecolor="orange", edgecolor="black", label="Near knobs"),
            Patch(facecolor="lightcoral", edgecolor="black", label="Negative binders")]
    ax.legend(
        handles=legend_elements + highlight_legend,
        title="Legend",
        loc="upper right",
        fontsize=8)

    # --- save figure ---
    fig_path = output_dir / f"{output_prefix}_cysteine_clustering_dendrogram.png"
    plt.savefig(fig_path, dpi=300)
    plt.close()
    print(f"Dendrogram written to: {fig_path}")

    # ==========================================================
    # SECOND DENDROGRAM WITH GLOBAL DISPLAY IDS
    # ==========================================================

    # --- build seq_id -> display_id lookup ---
    seq_to_display_id = {}

    for cluster_data in clusters.values():
        seq_to_display_id.update(cluster_data["display_ids"])

    # keep non-seq labels unchanged (e.g. knobs)
    display_labels = []

    for seq_id in ids:

        if seq_id in seq_to_display_id:
            display_labels.append(
                f"seq {seq_to_display_id[seq_id]}")
        else:
            display_labels.append(
                f"knob {seq_id}")

    # --- second dendrogram ---
    plt.figure(figsize=(20, 10))
    plt.subplots_adjust(bottom=0.1)

    ddata_display = dendrogram(
        Z,
        labels=display_labels,
        leaf_rotation=90,
        leaf_font_size=4,
        color_threshold=0,
        above_threshold_color="black",
    )

    ax = plt.gca()

    # --- background cluster shading ---
    leaf_order = ddata_display["leaves"]

    ordered_clusters = [best_labels[i] for i in leaf_order]

    blocks = []
    start = 0

    for i in range(1, len(ordered_clusters)):

        if ordered_clusters[i] != ordered_clusters[i - 1]:
            blocks.append(
                (start, i - 1, ordered_clusters[i - 1])
            )

            start = i

    blocks.append(
        (start, len(ordered_clusters) - 1, ordered_clusters[-1])
    )

    for start, end, cluster in blocks:
        x_start = start * 10
        x_end = (end + 1) * 10

        ax.axvspan(
            x_start,
            x_end,
            color=cluster_to_color[cluster],
            alpha=0.08,
            zorder=0,
        )

    plt.title(
        "Hierarchical clustering based on cysteine patterns (display IDs)",
        fontsize=22,
        fontweight="bold",
    )

    plt.xlabel("Display IDs", fontsize=16)
    plt.ylabel("Manhattan distance", fontsize=16)

    # --- identical label coloring ---
    for label, leaf_idx in zip(
            ax.get_xmajorticklabels(),
            ddata_display["leaves"],
    ):

        seq_id = ids[leaf_idx]

        if seq_id in knob_ids:

            label.set_fontweight("bold")

            label.set_bbox(
                dict(
                    facecolor="lightgray",
                    edgecolor="none",
                    boxstyle="round,pad=0.15",
                )
            )

        elif seq_id in negative_binder_id_set:

            label.set_bbox(
                dict(
                    facecolor="lightcoral",
                    edgecolor="none",
                    boxstyle="round,pad=0.15",
                )
            )

        elif seq_id in near_knob_ids:

            label.set_bbox(
                dict(
                    facecolor="orange",
                    edgecolor="none",
                    boxstyle="round,pad=0.15",
                )
            )

    ax.set_ylim(0, Z[:, 2].max() * 1.05)

    ax.legend(
        handles=legend_elements + highlight_legend,
        title="Legend",
        loc="upper right",
        fontsize=8,
    )

    fig_path_display = (
            output_dir /
            f"{output_prefix}_cysteine_clustering_dendrogram_display_ids.png"
    )

    plt.savefig(fig_path_display, dpi=300)

    plt.close()

    print(f"Dendrogram with display IDs written to: {fig_path_display}")

    if debug:
        print("\n=== Cysteine clustering END ===")

    return clusters, Z, ids, near_knob_to_knobs, knob_ids


def plot_cluster_logos(clusters: dict, save_dir: Path, output_prefix: str, filename: str | None = None, include_gaps: bool = False, highlight_aa: str | None = None
) -> None:
    """
    Plot sequence logos for each cluster in a stacked multi-panel figure.
    """
    if filename is None:
        gap_part = "with_gaps" if include_gaps else "no_gaps"

        if highlight_aa is None:
            color_part = "chemistry"
        else:
            color_part = f"highlight_{highlight_aa}"

        filename = (
            f"{output_prefix}_cluster_logos_"
            f"{gap_part}_{color_part}.png")

    n_clusters = len(clusters)
    total_sequences = sum(len(c["sequences"]) for c in clusters.values())

    fig, axes = plt.subplots(
        n_clusters, 1,
        figsize=(12, 3 * n_clusters),
        sharex=True)

    if n_clusters == 1:
        axes = [axes]

    for ax, (cluster_label, data) in zip(axes, sorted(clusters.items())):
        sequences = data["sequences"]
        n_seq_cluster = len(sequences)

        counts = logomaker.alignment_to_matrix(
            sequences,
            to_type="counts",
            characters_to_ignore="-")
        counts = counts.loc[counts.sum(axis=1) > 0]

        if include_gaps:
            freq = counts / n_seq_cluster
        else:
            freq = counts.div(counts.sum(axis=1), axis=0)

        if highlight_aa is None:
            logo = logomaker.Logo(
                freq,
                ax=ax,
                color_scheme="chemistry")
        else:
            color_dict = {
                aa: ("gold" if aa == highlight_aa else "lightgray")
                for aa in freq.columns}

            logo = logomaker.Logo(
                freq,
                ax=ax,
                color_scheme=color_dict)

        ax.set_ylabel("Frequency", fontsize=14)
        ax.set_title(
            f"Cluster {cluster_label + 1} (n = {n_seq_cluster})",
            fontsize=15)
        axes[-1].set_xlabel("Alignment position", fontsize=14)

    fig.suptitle(
        f"Cysteine-topology clusters "
        f"(n = {total_sequences} sequences, {n_clusters} clusters)",
        fontsize=18,
        y=0.98)

    plt.tight_layout(rect=[0, 0, 1, 0.96])

    out_path = save_dir / filename
    plt.savefig(out_path, dpi=300)
    plt.close()

    print(f"Cluster logos written to: {out_path} ")


def consensus_sequence(sequences: list[str]) -> str:
    """
    Build a gap-aware consensus sequence:
    most frequent amino acid per position.
    """
    max_len = max(len(seq) for seq in sequences)
    consensus = []

    for i in range(max_len):
        column = [
            seq[i] for seq in sequences
            if i < len(seq) and seq[i] != "-"
        ]
        if column:
            aa, _ = Counter(column).most_common(1)[0]
            consensus.append(aa)
        else:
            consensus.append("-")

    return "".join(consensus)


def cysteine_positions_and_spacing(seq: str):
    """
    Extract cysteine positions (1-based) and C–C spacings.
    """
    c_positions = [i + 1 for i, aa in enumerate(seq) if aa == "C"]
    c_spacings = [
        c_positions[i + 1] - c_positions[i] - 1
        for i in range(len(c_positions) - 1)
    ]
    return c_positions, c_spacings


def cluster_summary_to_latex( clusters: dict, output_tex_path
) -> None:
    """
    Generate a LaTeX table summarizing clusters:
    - cluster id
    - number of sequences
    - consensus sequence
    - cysteine count
    - cysteine positions
    - C–C spacing
    """
    rows = []
    for cluster_id, data in sorted(clusters.items()):
        sequences = data["sequences"]
        cons = consensus_sequence(sequences)
        c_pos, c_spacing = cysteine_positions_and_spacing(cons)

        rows.append({
            "Cluster": cluster_id + 1,
            "n": len(sequences),
            "Consensus": cons,
            "Cys count": len(c_pos),
            "Cys positions": ", ".join(map(str, c_pos)),
            "C–C spacing": ", ".join(map(str, c_spacing)) if c_spacing else "-"
        })

    df = pd.DataFrame(rows)
    print("\nCluster summary (preview):")
    print(df.to_markdown(index=False))
    df.to_latex(
        output_tex_path,
        index=False,
        escape=True,
        column_format="rcllll")
    print(f"Cluster summary table written to: {output_tex_path}")


def plot_cysteine_position_heatmap(clusters: dict, save_path, max_positions: int | None = None
) -> None:
    """
    Plot a heatmap showing cysteine frequency per alignment position per cluster.
    """
    # determine max sequence length
    if max_positions is None:
        max_positions = max(
            max(len(seq) for seq in data["sequences"])
            for data in clusters.values()
        )
    heatmap = []
    for cluster_id, data in sorted(clusters.items()):
        seqs = data["sequences"]
        n = len(seqs)

        freq = np.zeros(max_positions)

        for seq in seqs:
            for i, aa in enumerate(seq):
                if aa == "C":
                    freq[i] += 1

        freq = freq / n  # normalize
        heatmap.append(freq)
    heatmap = np.array(heatmap)

    # create labels with sequence counts
    cluster_labels = [
        f"Cluster {i + 1} (n={len(data['sequences'])})"
        for i, (_, data) in enumerate(sorted(clusters.items()))]
    n_clusters = len(clusters)

    # --- adaptive fontsize ---
    base_size = 12
    ylabel_size = max(10, 16 - n_clusters * 0.2)  # skaliert mit Clusterzahl
    title_size = 18
    tick_size = max(9, 13 - n_clusters * 0.15)

    plt.figure(figsize=(14, 4 + n_clusters * 0.5))  # mehr Höhe!
    plt.imshow(heatmap, aspect="auto", cmap="viridis")

    # --- colorbar ---
    cbar = plt.colorbar()
    cbar.set_label("Cysteine frequency", fontsize=15)
    cbar.ax.tick_params(labelsize=11)

    # --- axes ---
    ax = plt.gca()
    plt.yticks(
        range(len(cluster_labels)),
        cluster_labels,
        fontsize=ylabel_size)
    plt.xticks(fontsize=tick_size)

    plt.xlabel("Alignment position", fontsize=16)
    plt.ylabel("Cluster", fontsize=16)
    plt.title(
        "Cysteine position frequency across clusters",
        fontsize=title_size)

    # highlight y label
    for i, (cluster_id, data) in enumerate(sorted(clusters.items())):
        n_seq = len(data["sequences"])
        if n_seq >= 10:
            label = ax.get_yticklabels()[i]
            label.set_fontweight("bold")
            label.set_bbox(dict(
                facecolor="darkorange",
                edgecolor="none",
                boxstyle="round,pad=0.2"))

    # Highlight major clusters with side markers (no data overlap)
    for i, (cluster_id, data) in enumerate(sorted(clusters.items())):
        n_seq = len(data["sequences"])
        if n_seq >= 10:
            ax.plot(
                -0.5, i,
                marker=">",
                markersize=12,
                color="darkorange",
                clip_on=False)

    plt.tight_layout()
    plt.savefig(save_path, dpi=300)
    plt.close()
    print(f"Saved cysteine position heatmap to {save_path}")


def plot_cysteine_spacing_violin(clusters: dict, save_path) -> None:
    """
    Plot violin plots of C–C spacer lengths per cluster.
    """

    cluster_list = list(sorted(clusters.items()))

    spacings = []
    sizes = []

    for cluster_id, data in cluster_list:
        cluster_spacings = []

        for seq in data["sequences"]:
            c_positions = [i for i, aa in enumerate(seq) if aa == "C"]
            for i in range(len(c_positions) - 1):
                cluster_spacings.append(
                    c_positions[i + 1] - c_positions[i] - 1
                )

        if not cluster_spacings:
            cluster_spacings = [0]

        spacings.append(cluster_spacings)
        sizes.append(len(data["sequences"]))

    plt.figure(figsize=(10, 5))
    parts = plt.violinplot(spacings, showmeans=True, showmedians=True)

    labels = [
        f"Cluster {i + 1} (n={sizes[i]})"
        for i in range(len(sizes))
    ]

    plt.xticks(range(1, len(labels) + 1), labels, rotation=90, fontsize=11)
    plt.ylabel("C–C spacing (amino acids)", fontsize=14)
    plt.title("Distribution of inter-cysteine spacings per cluster", fontsize=16)

    # Highlighting
    for i, size in enumerate(sizes):
        if size >= 10:
            parts['bodies'][i].set_facecolor("orange")
            parts['bodies'][i].set_edgecolor("black")
            parts['bodies'][i].set_alpha(0.9)
        else:
            parts['bodies'][i].set_alpha(0.4)

    plt.savefig(save_path, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"Saved cysteine spacer violin plot to {save_path}")



def compute_sequence_identity_matrix(alignment_path, near_knobs_to_knobs, debug=False
):
    """
    Compute sequence identity between near-knobs and their associated knobs.

    Uses alignment and only counts positions where BOTH sequences
    have amino acids (no gaps).

    Returns:
        dict:
        {
            near_id: {
                knob_id: {
                    "identity": float,
                    "coverage": float,
                    "matches": int,
                    "positions": int
                }
            }
        }
    """

    # load alignment
    alignment = AlignIO.read(alignment_path, "clustal")
    aln_dict = {rec.id: str(rec.seq) for rec in alignment}

    results = {}

    for near_id, knob_list in near_knobs_to_knobs.items():

        if near_id not in aln_dict:
            continue

        results[near_id] = {}
        near_seq = aln_dict[near_id]

        for knob_id in knob_list:

            if knob_id not in aln_dict:
                continue

            knob_seq = aln_dict[knob_id]

            matches = 0
            valid_positions = 0
            alignment_length = len(near_seq)

            for a, b in zip(near_seq, knob_seq):

                # ignore gaps
                if a == "-" or b == "-":
                    continue

                valid_positions += 1

                if a == b:
                    matches += 1

            if valid_positions == 0:
                identity = 0.0
            else:
                identity = matches / valid_positions

            coverage = valid_positions / alignment_length

            results[near_id][knob_id] = {
                "identity": identity,
                "coverage": coverage,
                "matches": matches,
                "positions": valid_positions,
            }

            if debug:
                print(
                    f"\n[{near_id} vs {knob_id}] "
                    f"identity={identity:.3f} "
                    f"coverage={coverage:.3f} "
                    f"(matches={matches}/{valid_positions})")

                match_line = []
                for a, b in zip(near_seq, knob_seq):
                    if a == "-" or b == "-":
                        match_line.append(" ")  # gap positions
                    elif a == b:
                        match_line.append("|")  # match
                    else:
                        match_line.append(".")  # mismatch

                prefix = "NEAR: "

                print(prefix + near_seq)
                print(" " * len(prefix) + "".join(match_line))
                print("KNOB: " + knob_seq)

    return results


def extract_pdb(name: str) -> str:
    """
    Extract valid 4-character PDB code.
    Must start with a digit (strict PDB rule).
    """

    parts = name.split("_")

    for p in parts:
        # remove suffix like ".1"
        cleaned = p.split(".")[0]

        # VALID PDB: first char digit + 3 alphanumeric
        if re.fullmatch(r"[0-9][A-Za-z0-9]{3}", cleaned):
            return cleaned.upper()

    return "NA"


def build_alignment_lookup(aln_path: Path):
    """Prepare alignment sequences for matching."""
    alignment = AlignIO.read(aln_path, "clustal")

    records = []
    for rec in alignment:
        seq = str(rec.seq).replace("-", "")
        records.append((rec.id, seq))

    return records


def match_knob_to_pdb(knob_seq: str, aln_records):
    """Match knob sequence to alignment and extract PDB."""
    knob_seq = knob_seq.replace("-", "")

    for name, aln_seq in aln_records:
        if knob_seq == aln_seq or knob_seq in aln_seq or aln_seq in knob_seq:
            return extract_pdb(name)

    return "NA"


def export_full_knob_excel(near_knobs_to_knobs: dict, clusters: dict, aln_path_knobs: Path, output_excel_path: Path,
                           negative_binder_ids: dict | None = None, knob_ids: set | None = None, identity_results: dict | None = None
):
    """
    Export Excel with:
    - All sequences except knobs
    - near_knobs (green)
    - negative binders (red, priority over near)
    - mapping to knobs
    - original negative binder names
    - knobs in separate sheet
    """

    # --- prepare alignment ---
    aln_records = build_alignment_lookup(aln_path_knobs)

    # --- sequence lookup ---
    seq_lookup = {}
    for c in clusters.values():
        for sid, seq in zip(c["ids"], c["sequences"]):
            seq_lookup[sid] = seq

    # --- sets ---
    if knob_ids is None:
        raise ValueError("knob_ids must be provided")

    near_ids = set(near_knobs_to_knobs.keys())
    negative_ids = set(negative_binder_ids.keys()) if negative_binder_ids else set()

    # ============================
    # MAIN SHEET
    # ============================
    rows_main = []

    for seq_id, seq in seq_lookup.items():
        identity_str = ""

        # skip knobs
        if seq_id in knob_ids:
            continue

        seq_nogap = seq.replace("-", "")

        # --- classification (FIXED PRIORITY) ---
        if seq_id in negative_ids:
            type_label = "negative"
            related = ""
            original_name = negative_binder_ids[seq_id]["source"]["name"]


        elif seq_id in near_ids:
            type_label = "near_knob"
            related_knobs = near_knobs_to_knobs.get(seq_id, [])
            related = ",".join(related_knobs)
            original_name = ""

            # --- NEW: sequence identity ---
            if identity_results and seq_id in identity_results:
                identities = []

                for k in related_knobs:
                    val = identity_results[seq_id].get(k)
                    if val:
                        identities.append(f"{val['identity']:.2f}")
                    else:
                        identities.append("nan")
                identity_str = ",".join(identities)
            else:
                identity_str = ""
        else:
            type_label = "normal"
            related = ""
            original_name = ""

        rows_main.append({
            "ID": seq_id,
            "Sequence_no_gaps": seq_nogap,
            "Type": type_label,
            "Related_Knobs": related,
            "Sequence_identity_to_knobs": identity_str if seq_id in near_ids else "",
            "Original_Name": original_name})

    df_main = pd.DataFrame(rows_main)

    # ============================
    # KNOB SHEET
    # ============================
    knob_rows = []

    for knob_id in knob_ids:
        knob_seq = seq_lookup.get(knob_id, "")
        knob_seq_nogap = knob_seq.replace("-", "")

        pdb = match_knob_to_pdb(knob_seq, aln_records)

        knob_rows.append({
            "Knob_ID": knob_id,
            "Sequence_no_gaps": knob_seq_nogap,
            "PDB": pdb,
        })

    df_knob = pd.DataFrame(knob_rows)

    # ============================
    # WRITE EXCEL
    # ============================
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        df_main.to_excel(writer, sheet_name="Sequences", index=False)
        df_knob.to_excel(writer, sheet_name="Knobs", index=False)

    # ============================
    # COLOR FORMATTING
    # ============================
    wb = load_workbook(output_excel_path)
    ws = wb["Sequences"]

    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # find Type column
    type_col = None
    for i, cell in enumerate(ws[1]):
        if cell.value == "Type":
            type_col = i + 1

    for row in ws.iter_rows(min_row=2):
        cell = row[type_col - 1]

        if cell.value == "negative":
            for c in row:
                c.fill = red_fill

        elif cell.value == "near_knob":
            for c in row:
                c.fill = green_fill

    wb.save(output_excel_path)

    print(f"Extended Excel written to: {output_excel_path}")
